"""
Ingest labelled examples from an Excel file.

Expected sheet layout (the script tries to auto-detect columns by header name):
    - one column with the user's smeta description (header contains: 'наименование', 'description', 'запрос', 'позиция')
    - one column with the matched KSR code     (header contains: 'код', 'code', 'шифр')

Usage:
    python ingest_xlsx.py "C:\\path\\to\\smeta.xlsx"
    python ingest_xlsx.py "C:\\path\\to\\smeta.xlsx" --sheet Лист1 --query-col B --code-col D --source "smeta-vodopad"

Any row whose code is not found in the items table is reported in the skipped list.
"""
import argparse
import os
import sqlite3
import sys
from datetime import datetime, timezone

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(HERE, "ksr.sqlite")

QUERY_HEADERS = ("наименование", "description", "запрос", "позиция", "name", "описание")
CODE_HEADERS = ("код", "code", "шифр", "ксрр", "ксрп")


def _norm(s: str) -> str:
    import re
    if not s:
        return ""
    s = s.lower().replace("ё", "е")
    s = re.sub(r"[^a-zа-я0-9.,х/\-+\s]", " ", s)
    s = re.sub(r"(\d+),(\d+)", r"\1.\2", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _find_col(headers: list, candidates: tuple) -> int | None:
    for i, h in enumerate(headers):
        if not h:
            continue
        hl = str(h).strip().lower()
        for cand in candidates:
            if cand in hl:
                return i
    return None


def main():
    p = argparse.ArgumentParser()
    p.add_argument("xlsx", help="Path to labelled .xlsx file")
    p.add_argument("--sheet", help="Sheet name (default: first sheet)")
    p.add_argument("--query-col", help="Column letter for query/description (overrides auto-detect)")
    p.add_argument("--code-col", help="Column letter for KSR code (overrides auto-detect)")
    p.add_argument("--header-row", type=int, default=1, help="Header row number (1-based, default 1)")
    p.add_argument("--source", default=None, help="Source tag for stored examples (default: file:<basename>)")
    p.add_argument("--weight", type=float, default=1.0, help="Weight of each example (default 1.0)")
    p.add_argument("--dry-run", action="store_true", help="Show what would be inserted, do not write")
    args = p.parse_args()

    if not os.path.exists(args.xlsx):
        print(f"File not found: {args.xlsx}"); sys.exit(1)
    if not os.path.exists(DB_PATH):
        print(f"DB not found: {DB_PATH}\nRun build_index.py first."); sys.exit(1)

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Empty sheet"); sys.exit(1)

    headers = [str(c).strip() if c is not None else "" for c in rows[args.header_row - 1]]

    def _col_idx(letter: str | None, candidates: tuple) -> int | None:
        if letter:
            return openpyxl.utils.column_index_from_string(letter.strip()) - 1
        return _find_col(headers, candidates)

    qi = _col_idx(args.query_col, QUERY_HEADERS)
    ci = _col_idx(args.code_col, CODE_HEADERS)

    if qi is None or ci is None:
        print(f"Couldn't auto-detect columns.\n  Headers found: {headers}")
        print(f"  query col: {qi} (need one of {QUERY_HEADERS})")
        print(f"  code col:  {ci} (need one of {CODE_HEADERS})")
        print("  Pass --query-col X --code-col Y explicitly.")
        sys.exit(2)

    print(f"Sheet: {ws.title} | query col: {chr(65+qi)} | code col: {chr(65+ci)}")

    source = args.source or f"file:{os.path.basename(args.xlsx)}"

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS examples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            query_text TEXT NOT NULL,
            query_norm TEXT NOT NULL,
            item_id INTEGER NOT NULL,
            source TEXT NOT NULL,
            weight REAL DEFAULT 1.0,
            added_at TEXT NOT NULL
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_examples_norm ON examples(query_norm)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_examples_item ON examples(item_id)")

    added, skipped = 0, []
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    for ri, row in enumerate(rows[args.header_row:], start=args.header_row + 1):
        if qi >= len(row) or ci >= len(row):
            continue
        q = row[qi]
        code = row[ci]
        if q is None or code is None:
            continue
        q = str(q).strip()
        code = str(code).strip()
        if not q or not code:
            continue
        cur.execute("SELECT id FROM items WHERE code=? LIMIT 1", (code,))
        r = cur.fetchone()
        if not r:
            skipped.append((ri, code, q[:60]))
            continue
        if args.dry_run:
            print(f"  [{ri}] {code:<22} ← {q[:60]}")
        else:
            cur.execute(
                "INSERT INTO examples(query_text, query_norm, item_id, source, weight, added_at) "
                "VALUES(?,?,?,?,?,?)",
                (q, _norm(q), r[0], source, args.weight, now),
            )
        added += 1

    if not args.dry_run:
        conn.commit()
    conn.close()

    print(f"\n{'DRY-RUN: would add' if args.dry_run else 'Added'}: {added}")
    print(f"Skipped (code not in classifier): {len(skipped)}")
    for ri, code, q in skipped[:20]:
        print(f"  row {ri}: code={code!r}  query={q!r}")
    if len(skipped) > 20:
        print(f"  ... and {len(skipped) - 20} more")


if __name__ == "__main__":
    main()
