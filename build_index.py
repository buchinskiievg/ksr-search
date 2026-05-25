"""
Build SQLite + TF-IDF index from KSR classifier Excel file.
Run once:  python build_index.py "C:\path\to\Классификатор 13022026.xlsx"
"""
import sys
import os
import re
import sqlite3
import pickle
import openpyxl
from scipy.sparse import hstack, save_npz
from sklearn.feature_extraction.text import TfidfVectorizer

HERE = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(HERE, "ksr.sqlite")
INDEX_PATH = os.path.join(HERE, "ksr_index.pkl")

CODE_RE = re.compile(r"^\d{2}(\.\d+)+(-\d+)?(-\d+)?$|^\d{2}\.\d+\.\d+\.\d+\.\d+\.\d+\.\d+\.\d+-\d+-\d+$")
HEADER_PREFIX = ("Книга ", "Часть ", "Раздел ", "Группа ", "Подгруппа ")


def normalize(s: str) -> str:
    if not s:
        return ""
    s = s.lower().replace("ё", "е")
    s = re.sub(r"[^a-zа-я0-9.,х/\-+\s]", " ", s)
    s = re.sub(r"(\d+),(\d+)", r"\1.\2", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def read_xlsx(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        path_stack = {"Книга": "", "Часть": "", "Раздел": "", "Группа": "", "Подгруппа": ""}
        for row in ws.iter_rows(min_row=3, values_only=True):
            code, name, unit = (row + (None, None, None))[:3]
            if code is None and name is None:
                continue
            code = str(code).strip() if code is not None else ""
            name = str(name).strip() if name is not None else ""
            unit = str(unit).strip() if unit is not None else ""
            if not name:
                continue
            # Header rows have same value in all 3 cells and start with header keyword
            if code == name and any(name.startswith(p) for p in HEADER_PREFIX):
                for key in path_stack:
                    if name.startswith(key + " "):
                        # reset deeper levels
                        levels = ["Книга", "Часть", "Раздел", "Группа", "Подгруппа"]
                        idx = levels.index(key)
                        path_stack[key] = name
                        for deeper in levels[idx + 1:]:
                            path_stack[deeper] = ""
                        break
                continue
            category = " / ".join(v for v in path_stack.values() if v)
            items.append({
                "sheet": sheet_name,
                "code": code,
                "name": name,
                "unit": unit,
                "category": category,
            })
    return items


def main():
    if len(sys.argv) < 2:
        print("Usage: python build_index.py <path-to-xlsx>")
        sys.exit(1)
    src = sys.argv[1]
    if not os.path.exists(src):
        print(f"File not found: {src}")
        sys.exit(1)

    print(f"[1/4] Reading {src} ...")
    items = read_xlsx(src)
    print(f"      {len(items):,} item rows")

    print("[2/4] Writing SQLite ...")
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE items(
        id INTEGER PRIMARY KEY,
        sheet TEXT, code TEXT, name TEXT, unit TEXT, category TEXT
    )""")
    c.executemany(
        "INSERT INTO items(id, sheet, code, name, unit, category) VALUES(?,?,?,?,?,?)",
        [(i, it["sheet"], it["code"], it["name"], it["unit"], it["category"])
         for i, it in enumerate(items)]
    )
    c.execute("CREATE INDEX idx_code ON items(code)")
    conn.commit()
    conn.close()

    print("[3/4] Building TF-IDF (word + char n-grams) ...")
    texts = [normalize(it["name"]) for it in items]

    word_vec = TfidfVectorizer(
        analyzer="word", ngram_range=(1, 2),
        min_df=2, max_df=0.9, sublinear_tf=True, max_features=200_000,
    )
    char_vec = TfidfVectorizer(
        analyzer="char_wb", ngram_range=(3, 5),
        min_df=2, max_df=0.95, sublinear_tf=True, max_features=300_000,
    )
    word_mat = word_vec.fit_transform(texts)
    char_mat = char_vec.fit_transform(texts)
    print(f"      word vocab: {len(word_vec.vocabulary_):,}, char vocab: {len(char_vec.vocabulary_):,}")

    print("[4/4] Saving index ...")
    with open(INDEX_PATH, "wb") as f:
        pickle.dump({
            "word_vec": word_vec,
            "char_vec": char_vec,
            "word_mat": word_mat,
            "char_mat": char_mat,
            "n_items": len(items),
        }, f)
    print(f"Done. {DB_PATH}, {INDEX_PATH}")


if __name__ == "__main__":
    main()
